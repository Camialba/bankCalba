
from openpyxl import Workbook, load_workbook

class DBExcel:

    def abrirExcel (self):
        try:
            self.exc = load_workbook("cuentasbancarias.xlsx")
        except FileNotFoundError:
            print ("Archivo inexistente")
            self.crearArchivo()

    def crearArchivo (self):
        
        self.exc = Workbook()

        #trabaja con la hoja activa que se crea por def

        self.hoja = self.exc.active
        self.hoja.title = "Usuarios"

        self.hoja["A1"] = "DNI"
        self.hoja["B1"] = "Nombre titular"
        self.hoja["C1"] = "Tipo de cuenta"

        #se crea una nueva hoja

        self.hoja2 = self.exc.create_sheet(title= "Movimientos")
        
        self.hoja2["A1"] = "DNI"
        self.hoja2["B1"] = "Tipo movimiento"
        self.hoja2["C1"] = "Monto"


        self.exc.save("cuentasbancarias.xlsx")

    def registrar(self, pagina, datos):
        self.abrirExcel()
        self.hojaRegistrada = self.exc[pagina]
        self.hojaRegistrada.append(datos)
        self.exc.save("cuentasbancarias.xlsx")

    def registrarCliente(self, datosCliente):
        self.registrar("Usuarios", datosCliente)

    def registrarDep(self, datosDep):
        datos = list(datosDep)
        datos.insert (1, "Deposito")
        self.registrar("Movimientos", datos)

    def registrarExtr(self, datosExtr):
        datosE = list(datosExtr)
        datosE.insert (1, "Extracción")
        self.registrar("Movimientos", datosE)

#prueba
app = DBExcel()
# app.registrarDep((38851514, 500))
app.registrarExtr((38851514, 250))



